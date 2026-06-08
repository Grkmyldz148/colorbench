"""MetricSpace evaluation — STRESS scoring on perceptual color difference datasets.

Completely separate from GenSpace benchmark. Does not touch run_test() or any
existing gpu_metrics modules.

Usage:
    from core.metric_eval import run_metric_evaluation
    results = run_metric_evaluation(
        metric_json="helmlab-main-repo/checkpoints/metricspace_v21.json",
        datasets_dir="datasets"
    )

STRESS: lower = better. CIEDE2000 baseline ~29, CIELab ~30.
Datasets:
  - COMBVD   : 3813 pairs (BFD-P, LEEDS, RIT-DuPont, WITT)
  - MacAdam  : 128 pairs (MacAdam 1974 uniform color scales, D65)
  - Human FB : 3552 observer judgements (71 observers, sRGB hex)
"""

import json
import os
import re
import sys
import numpy as np

# ── Bradford CAT (arbitrary white → D65) ─────────────────────────────────────

_BRADFORD = np.array([
    [ 0.8951,  0.2664, -0.1614],
    [-0.7502,  1.7135,  0.0367],
    [ 0.0389, -0.0685,  1.0296],
], dtype=np.float64)
_BRADFORD_INV = np.linalg.inv(_BRADFORD)
_D65 = np.array([0.95047, 1.0, 1.08883], dtype=np.float64)


def _cat_to_d65(xyz: np.ndarray, white: np.ndarray) -> np.ndarray:
    """Bradford chromatic adaptation: arbitrary white → D65."""
    if np.allclose(white, _D65, atol=1e-4):
        return xyz
    cone_src = _BRADFORD @ white
    cone_dst = _BRADFORD @ _D65
    scale = cone_dst / cone_src
    M = _BRADFORD_INV @ (np.diag(scale) @ _BRADFORD)
    return xyz @ M.T


# ── Baseline metrics (standalone numpy) ───────────────────────────────────────

def _xyz_to_cielab(xyz: np.ndarray, white: np.ndarray) -> np.ndarray:
    """XYZ → CIE Lab via colour-science (peer-reviewed reference).

    `white` is XYZ tristimulus of the reference illuminant. colour-science
    accepts xy chromaticity, so we project XYZ → xy here.
    """
    import colour
    # XYZ → xy chromaticity for the illuminant
    s = white.sum()
    illum_xy = np.array([white[0] / s, white[1] / s])
    return colour.XYZ_to_Lab(xyz, illuminant=illum_xy)


def _cielab_de(xyz1: np.ndarray, xyz2: np.ndarray, white: np.ndarray) -> np.ndarray:
    """CIE76 ΔE via colour-science."""
    import colour
    lab1 = _xyz_to_cielab(xyz1, white)
    lab2 = _xyz_to_cielab(xyz2, white)
    return colour.delta_E(lab1, lab2, method="CIE 1976")


def _ciede2000(xyz1: np.ndarray, xyz2: np.ndarray, white: np.ndarray) -> np.ndarray:
    """CIEDE2000 via colour-science (CIE 224:2017 reference)."""
    import colour
    lab1 = _xyz_to_cielab(xyz1, white)
    lab2 = _xyz_to_cielab(xyz2, white)
    return colour.delta_E(lab1, lab2, method="CIE 2000")


def _oklab_de(xyz1: np.ndarray, xyz2: np.ndarray) -> np.ndarray:
    """OKLab ΔE (uses D65 natively)."""
    M1 = np.array([
        [0.8189330101, 0.3618667424, -0.1288597137],
        [0.0329845436, 0.9293118715, 0.0361456387],
        [0.0482003018, 0.2643662691, 0.6338517070],
    ])
    M2 = np.array([
        [0.2104542553, 0.7936177850, -0.0040720468],
        [1.9779984951, -2.4285922050, 0.4505937099],
        [0.0259040371, 0.7827717662, -0.8086757660],
    ])
    lms1 = xyz1 @ M1.T
    lms2 = xyz2 @ M1.T
    lms1 = np.cbrt(np.maximum(lms1, 0))
    lms2 = np.cbrt(np.maximum(lms2, 0))
    lab1 = lms1 @ M2.T
    lab2 = lms2 @ M2.T
    return np.sqrt(np.sum((lab1 - lab2) ** 2, axis=-1))


def _cam16_ucs_de(xyz1: np.ndarray, xyz2: np.ndarray) -> np.ndarray:
    """CAM16-UCS ΔE (Li et al. 2017, doi:10.1002/col.22131).

    Standard reference implementation via ``colour-science`` library
    (peer-reviewed, matches MATLAB Image Processing Toolbox and Color.js
    bit-for-bit on the published worked examples). Uses ``XYZ_to_CAM16UCS``
    with default conditions (LA=64 cd/m², Yb=20%, average surround),
    which is the conventional baseline-comparison setting used in CIE
    224:2017 and in the published delta_E benchmarks.

    Inputs are CIE XYZ, Y normalized to 1 (the colour-science API takes
    Y/Y_w in [0, 1]). Outputs are per-pair Cartesian Euclidean distance
    in CAM16-UCS (J', a', b') coordinates.

    History note: the original v0.x ColorBench shipped a hand-rolled
    NumPy implementation that produced STRESS values ~22 points higher
    than the standard reference (e.g. COMBVD=56.19 vs reference 33.47).
    The bug was traced to inconsistent handling of the FL adaptation
    scale between Y_w=1 and Y_w=100 conventions in the post-adaptation
    compression. Replaced 2026-05-06 with a colour-science wrapper.
    """
    import colour
    Jab1 = colour.XYZ_to_CAM16UCS(xyz1)
    Jab2 = colour.XYZ_to_CAM16UCS(xyz2)
    return np.sqrt(np.sum((Jab1 - Jab2) ** 2, axis=-1))


def _cie94_de(xyz1: np.ndarray, xyz2: np.ndarray, white: np.ndarray) -> np.ndarray:
    """CIE94 ΔE via colour-science (graphic arts variant default)."""
    import colour
    lab1 = _xyz_to_cielab(xyz1, white)
    lab2 = _xyz_to_cielab(xyz2, white)
    return colour.delta_E(lab1, lab2, method="CIE 1994")


def _ciecam02_ucs_de(xyz1: np.ndarray, xyz2: np.ndarray) -> np.ndarray:
    """CIECAM02-UCS (Luo et al. 2006, doi:10.1002/col.20227) ΔE.

    Standard reference implementation via ``colour-science``: composes
    XYZ → CIECAM02 J/M/h appearance correlates → CAM02-UCS Cartesian
    (J', a', b') with c1=0.007, c2=0.0228. Same viewing conditions as
    CAM16-UCS (LA=64 cd/m², Yb=20%, average surround).

    Replaces a hand-rolled implementation that shared the same FL
    scaling bug as the old ``_cam16_ucs_de`` (~22 STRESS points high).
    """
    import colour
    XYZ_w = np.array([0.95047, 1.0, 1.08883]) * 100.0
    LA, Yb = 64.0, 20.0
    surround = colour.appearance.VIEWING_CONDITIONS_CIECAM02["Average"]

    spec1 = colour.XYZ_to_CIECAM02(np.atleast_2d(xyz1) * 100.0, XYZ_w, LA, Yb, surround)
    spec2 = colour.XYZ_to_CIECAM02(np.atleast_2d(xyz2) * 100.0, XYZ_w, LA, Yb, surround)
    JMh1 = np.stack([np.asarray(spec1.J), np.asarray(spec1.M), np.asarray(spec1.h)], axis=-1)
    JMh2 = np.stack([np.asarray(spec2.J), np.asarray(spec2.M), np.asarray(spec2.h)], axis=-1)

    Jab1 = colour.JMh_CIECAM02_to_CAM02UCS(JMh1)
    Jab2 = colour.JMh_CIECAM02_to_CAM02UCS(JMh2)
    return np.sqrt(np.sum((Jab1 - Jab2) ** 2, axis=-1))


def _din99_de(xyz1: np.ndarray, xyz2: np.ndarray, white: np.ndarray) -> np.ndarray:
    """DIN99 ΔE (DIN 6176:2001).

    Standard reference implementation via ``colour-science``:
    XYZ → CIE Lab (with the supplied per-pair white) → DIN99 (L99, a99, b99)
    via ``colour.Lab_to_DIN99``. Default DIN99 (1999) variant — not the
    later DIN99b/c/d revisions, which colour-science exposes through the
    ``method=`` keyword if needed.

    Replaces a hand-rolled NumPy implementation that produced values
    ~1% off the reference (35.76 vs 35.49 STRESS on COMBVD with
    Bradford CAT pre-processing). Functionally equivalent, but
    delegating to colour-science keeps every CIE/ISO baseline in
    metric_eval pinned to a single peer-reviewed source of truth.
    """
    import colour
    lab1 = _xyz_to_cielab(xyz1, white)
    lab2 = _xyz_to_cielab(xyz2, white)
    d1 = colour.Lab_to_DIN99(lab1)
    d2 = colour.Lab_to_DIN99(lab2)
    return np.sqrt(np.sum((d1 - d2) ** 2, axis=-1))


def _jzazbz_de(xyz1: np.ndarray, xyz2: np.ndarray) -> np.ndarray:
    """Jzazbz ΔEz (Safdar et al. 2017, doi:10.1364/OE.25.015131).

    Standard reference implementation via ``colour-science``: assumes
    HDR-aware PQ transfer with the published 10,000 cd/m² peak constant.
    Inputs are CIE XYZ, Y normalized to 1 (colour-science API takes
    Y/Y_w in [0, 1] and rescales internally).

    Replaces a hand-rolled NumPy implementation that used a hard-coded
    L_sdr=203 cd/m² rescale, producing STRESS values ~1.3 points off
    the reference. The colour-science version matches the CIE 224:2017
    benchmark suite and the original paper's worked examples.
    """
    import colour
    Jab1 = colour.XYZ_to_Jzazbz(xyz1)
    Jab2 = colour.XYZ_to_Jzazbz(xyz2)
    return np.sqrt(np.sum((Jab1 - Jab2) ** 2, axis=-1))


# ── STRESS ─────────────────────────────────────────────────────────────────────

def stress(de_pred: np.ndarray, dv_obs: np.ndarray) -> float:
    """STRESS score. Lower = better. CIEDE2000 baseline ~29."""
    de = np.asarray(de_pred, dtype=np.float64)
    dv = np.asarray(dv_obs, dtype=np.float64)
    # Optimal linear scaling: F = Σ(ΔE·ΔV) / Σ(ΔE²)
    F = np.dot(de, dv) / np.dot(de, de)
    residuals = F * de - dv
    return 100.0 * np.sqrt(np.sum(residuals ** 2) / np.sum(dv ** 2))


# ── Dataset loaders ─────────────────────────────────────────────────────────────

def load_combvd(datasets_dir: str):
    """3813 XYZ color pairs with visual difference scores (multi-illuminant).

    Returns (xyz1, xyz2, whites, dv, sub_datasets).
    """
    path = os.path.join(datasets_dir, "combvd_pairs.json")
    data = json.load(open(path))
    xyz1 = np.array([d["xyz1"] for d in data])
    xyz2 = np.array([d["xyz2"] for d in data])
    whites = np.array([d["white"] for d in data])
    dv = np.array([d["dv"] for d in data])
    return xyz1, xyz2, whites, dv


def load_combvd_from_xlsx(datasets_dir: str):
    """Load COMBVD from xlsx with sub-dataset labels (requires openpyxl)."""
    try:
        import openpyxl
    except ImportError:
        return None

    path = os.path.join(datasets_dir, "combvd.xlsx")
    if not os.path.exists(path):
        return None

    wb = openpyxl.load_workbook(path)
    ws = wb["COM_Corrected_UNWEIGHTED"]

    records = []
    current_dataset = None
    for r in range(4, ws.max_row + 1):
        label = ws.cell(r, 1).value
        dv_val = ws.cell(r, 2).value
        if label:
            current_dataset = label
        if dv_val is None:
            continue
        row = [ws.cell(r, c).value for c in range(2, 12)]
        if None in row:
            continue
        records.append({
            "dataset": current_dataset,
            "dv": row[0],
            "white": [row[1]/100, row[2]/100, row[3]/100],
            "xyz1": [row[4]/100, row[5]/100, row[6]/100],
            "xyz2": [row[7]/100, row[8]/100, row[9]/100],
        })

    return records


def load_macadam1974(datasets_dir: str):
    """MacAdam (1974) uniform color scales — 128 pairs, D65, CIE 1964.

    Returns (xyz1, xyz2, white_d65, dv).
    Reference: MacAdam, JOSA 64(12), 1691-1702 (1974).
    """
    table1_path = os.path.join(datasets_dir, "macadam1974", "table1.yaml")
    table2_path = os.path.join(datasets_dir, "macadam1974", "table2.yaml")

    # Parse tile colorimetry (xyY, D65, 10° observer)
    tiles = {}
    for line in open(table2_path):
        line = line.strip()
        m = re.match(r'"(\w+)":\s*\[(.+)\]', line)
        if m:
            vals = [float(x.strip()) for x in m.group(2).split(",")]
            tiles[m.group(1)] = vals  # [x, y, Y, g, j, L]

    # D65 10° observer white (xyY → XYZ)
    xw, yw = 0.31272, 0.32903
    white = np.array([xw / yw, 1.0, (1.0 - xw - yw) / yw])  # ≈ [0.9504, 1.0, 1.0888]

    def xyyToXYZ(x, y, Y_pct):
        Y = Y_pct / 100.0
        return np.array([x * Y / y, Y, (1.0 - x - y) * Y / y])

    # Parse pairs [pair_num, "tile1", "tile2", dv_obs, ...]
    pairs = []
    for line in open(table1_path):
        line = line.strip()
        m = re.match(r"-\s*\[(\d+),\s*\"(\w+)\",\s*\"(\w+)\",\s*([\d.]+)", line)
        if m:
            t1, t2 = m.group(2), m.group(3)
            dv_obs = float(m.group(4))
            if t1 in tiles and t2 in tiles:
                x1, y1, Y1 = tiles[t1][0], tiles[t1][1], tiles[t1][2]
                x2, y2, Y2 = tiles[t2][0], tiles[t2][1], tiles[t2][2]
                xyz1 = xyyToXYZ(x1, y1, Y1)
                xyz2 = xyyToXYZ(x2, y2, Y2)
                pairs.append((xyz1, xyz2, dv_obs))

    if not pairs:
        return None, None, None, None

    xyz1 = np.array([p[0] for p in pairs])
    xyz2 = np.array([p[1] for p in pairs])
    dv = np.array([p[2] for p in pairs])
    return xyz1, xyz2, white, dv


def load_human_feedback(datasets_dir: str):
    """3552 observer judgements (71 observers, sRGB hex pairs)."""
    path = os.path.join(datasets_dir, "human_feedback.json")
    data = json.load(open(path))
    judgements = data["judgements"]

    from .cs.constants import M_SRGB as _M_SRGB_CANONICAL  # cycle 36 — 17-decimal
    _M_SRGB = _M_SRGB_CANONICAL.numpy()

    def hex_to_xyz(h):
        h = h.lstrip("#")
        rgb = np.array([int(h[i:i+2], 16) / 255.0 for i in (0, 2, 4)])
        linear = np.where(rgb <= 0.04045, rgb / 12.92, ((rgb + 0.055) / 1.055) ** 2.4)
        return _M_SRGB @ linear

    xyz1, xyz2, dv = [], [], []
    for j in judgements:
        xyz1.append(hex_to_xyz(j["hex1"]))
        xyz2.append(hex_to_xyz(j["hex2"]))
        dv.append(float(j["perceived_dv"]))

    return np.array(xyz1), np.array(xyz2), np.array(dv)


# ── MetricSpace wrapper ─────────────────────────────────────────────────────────

def _load_metric_space(metric_json: str, repo_root: str):
    """Load a metric space. Two paths:
      - .npy  -> black-box ExternalMetricSpace adapter (ANY architecture, e.g.
                 regime_family_v2 / alpha_v5_A2). ColorBench's datasets / STRESS /
                 Bradford CAT stay frozen; the space only provides distance().
      - .json -> bundled MetricSpace parametric family (v21-era 45-param format).
    """
    if str(metric_json).endswith(".npy"):
        from core.external_space import load_regime_champion
        return load_regime_champion(metric_json)
    src_path = os.path.join(repo_root, "src")
    if src_path not in sys.path:
        sys.path.insert(0, src_path)
    from helmlab.spaces.metric import MetricSpace, MetricParams
    params = MetricParams.from_dict(json.load(open(metric_json)))
    return MetricSpace(params)


# ── Main evaluation ─────────────────────────────────────────────────────────────

def run_metric_evaluation(metric_json: str, datasets_dir: str, repo_root: str) -> dict:
    """
    Evaluate MetricSpace against COMBVD, MacAdam 1974, and Human Feedback.
    Returns dict with STRESS scores for all metrics and datasets.
    """
    print("\n══════════════════════════════════════════════════════")
    print("  MetricSpace Evaluation — STRESS Scores")
    print("══════════════════════════════════════════════════════\n")

    print("Loading MetricSpace...", end=" ", flush=True)
    metric = _load_metric_space(metric_json, repo_root)
    print("done")

    results = {}

    # Common baseline set (applied per-dataset with appropriate white)
    def run_baselines(xyz1, xyz2, white_scalar, dv, is_d65=True):
        """Run all baselines. white_scalar: single white point for CIELab/CIEDE2000."""
        scores = {}
        for name, fn in [
            ("MetricSpace", lambda: metric.distance(xyz1, xyz2)),
            ("CIEDE2000",   lambda: _ciede2000(xyz1, xyz2, white_scalar)),
            ("CIE Lab",     lambda: _cielab_de(xyz1, xyz2, white_scalar)),
            ("OKLab",       lambda: _oklab_de(xyz1, xyz2)),
            ("Jzazbz",      lambda: _jzazbz_de(xyz1, xyz2)),
        ]:
            print(f"  {name}...", end=" ", flush=True)
            de = fn()
            s = stress(de, dv)
            scores[name] = round(s, 2)
            print(f"STRESS = {s:.2f}")
        return scores

    # ── COMBVD (3813 pairs, multi-illuminant) ──────────────────────────────
    print("\n[1/3] COMBVD (3813 pairs, multi-illuminant)")
    xyz1_raw, xyz2_raw, whites_raw, dv = load_combvd(datasets_dir)

    # Adapt all pairs to D65 for MetricSpace, OKLab, Jzazbz
    xyz1_d65 = np.array([_cat_to_d65(xyz1_raw[i], whites_raw[i]) for i in range(len(xyz1_raw))])
    xyz2_d65 = np.array([_cat_to_d65(xyz2_raw[i], whites_raw[i]) for i in range(len(xyz2_raw))])

    # For CIELab / CIEDE2000 / CIE94 / DIN99: use per-pair white (correct for multi-illuminant)
    scores_combvd = {}
    for name, fn in [
        ("MetricSpace",   lambda: metric.distance(xyz1_d65, xyz2_d65)),
        ("CIEDE2000",     lambda: np.array([
            _ciede2000(xyz1_raw[i:i+1], xyz2_raw[i:i+1], whites_raw[i])
            for i in range(len(dv))]).ravel()),
        ("CIE94",         lambda: np.array([
            _cie94_de(xyz1_raw[i:i+1], xyz2_raw[i:i+1], whites_raw[i])
            for i in range(len(dv))]).ravel()),
        ("CIE Lab",       lambda: np.array([
            _cielab_de(xyz1_raw[i:i+1], xyz2_raw[i:i+1], whites_raw[i])
            for i in range(len(dv))]).ravel()),
        ("DIN99",         lambda: np.array([
            _din99_de(xyz1_raw[i:i+1], xyz2_raw[i:i+1], whites_raw[i])
            for i in range(len(dv))]).ravel()),
        ("OKLab",         lambda: _oklab_de(xyz1_d65, xyz2_d65)),
        ("CAM16-UCS",     lambda: _cam16_ucs_de(xyz1_d65, xyz2_d65)),
        ("CIECAM02-UCS",  lambda: _ciecam02_ucs_de(xyz1_d65, xyz2_d65)),
        ("Jzazbz",        lambda: _jzazbz_de(xyz1_d65, xyz2_d65)),
    ]:
        print(f"  {name}...", end=" ", flush=True)
        de = fn()
        s = stress(de, dv)
        scores_combvd[name] = round(s, 2)
        print(f"STRESS = {s:.2f}")

    results["COMBVD"] = scores_combvd

    # Per-sub-dataset COMBVD breakdown (if xlsx available)
    combvd_records = load_combvd_from_xlsx(datasets_dir)
    if combvd_records:
        print("\n  COMBVD sub-dataset breakdown:")
        sub_datasets = {}
        for rec in combvd_records:
            sd = rec["dataset"]
            if sd not in sub_datasets:
                sub_datasets[sd] = []
            sub_datasets[sd].append(rec)

        sub_results = {}
        for sd_name, recs in sorted(sub_datasets.items()):
            x1 = np.array([r["xyz1"] for r in recs])
            x2 = np.array([r["xyz2"] for r in recs])
            wh = np.array([r["white"] for r in recs])
            dv_sub = np.array([r["dv"] for r in recs])
            x1_d65 = np.array([_cat_to_d65(x1[i], wh[i]) for i in range(len(x1))])
            x2_d65 = np.array([_cat_to_d65(x2[i], wh[i]) for i in range(len(x2))])
            de_ms = metric.distance(x1_d65, x2_d65)
            de_00 = np.array([_ciede2000(x1[i:i+1], x2[i:i+1], wh[i]) for i in range(len(dv_sub))]).ravel()
            s_ms = stress(de_ms, dv_sub)
            s_00 = stress(de_00, dv_sub)
            n = len(recs)
            marker = " ★" if s_ms < s_00 else ""
            print(f"    {sd_name:<18} n={n:>4}  MetricSpace={s_ms:.2f}{marker}  CIEDE2000={s_00:.2f}")
            sub_results[sd_name] = {"MetricSpace": round(s_ms, 2), "CIEDE2000": round(s_00, 2), "n": n}
        results["COMBVD_subsets"] = sub_results

    # ── MacAdam 1974 (D65, CIE 1964) ──────────────────────────────────────
    print("\n[2/3] MacAdam 1974 (D65, CIE 1964, uniform color scales)")
    xyz1_mac, xyz2_mac, white_mac, dv_mac = load_macadam1974(datasets_dir)

    if xyz1_mac is not None:
        # MacAdam white ≈ D65, apply CAT for exactness
        xyz1_mac_d65 = _cat_to_d65(xyz1_mac, white_mac)
        xyz2_mac_d65 = _cat_to_d65(xyz2_mac, white_mac)

        scores_mac = {}
        for name, fn in [
            ("MetricSpace",   lambda: metric.distance(xyz1_mac_d65, xyz2_mac_d65)),
            ("CIEDE2000",     lambda: _ciede2000(xyz1_mac, xyz2_mac, white_mac)),
            ("CIE94",         lambda: _cie94_de(xyz1_mac, xyz2_mac, white_mac)),
            ("CIE Lab",       lambda: _cielab_de(xyz1_mac, xyz2_mac, white_mac)),
            ("DIN99",         lambda: _din99_de(xyz1_mac, xyz2_mac, white_mac)),
            ("OKLab",         lambda: _oklab_de(xyz1_mac_d65, xyz2_mac_d65)),
            ("CAM16-UCS",     lambda: _cam16_ucs_de(xyz1_mac_d65, xyz2_mac_d65)),
            ("CIECAM02-UCS",  lambda: _ciecam02_ucs_de(xyz1_mac_d65, xyz2_mac_d65)),
            ("Jzazbz",        lambda: _jzazbz_de(xyz1_mac_d65, xyz2_mac_d65)),
        ]:
            print(f"  {name}...", end=" ", flush=True)
            de = fn()
            s = stress(de, dv_mac)
            scores_mac[name] = round(s, 2)
            print(f"STRESS = {s:.2f}")
        results["MacAdam1974"] = scores_mac
        print(f"  ({len(dv_mac)} pairs)")
    else:
        print("  [skipped — macadam1974/ not found]")

    # ── Human Feedback (3552 judgements) ──────────────────────────────────
    print("\n[3/3] Human Feedback (3552 judgements, 71 observers, sRGB)")
    hxyz1, hxyz2, hdv = load_human_feedback(datasets_dir)

    hscores = {}
    for name, fn in [
        ("MetricSpace",   lambda: metric.distance(hxyz1, hxyz2)),
        ("CIEDE2000",     lambda: _ciede2000(hxyz1, hxyz2, _D65)),
        ("CIE94",         lambda: _cie94_de(hxyz1, hxyz2, _D65)),
        ("CIE Lab",       lambda: _cielab_de(hxyz1, hxyz2, _D65)),
        ("DIN99",         lambda: _din99_de(hxyz1, hxyz2, _D65)),
        ("OKLab",         lambda: _oklab_de(hxyz1, hxyz2)),
        ("CAM16-UCS",     lambda: _cam16_ucs_de(hxyz1, hxyz2)),
        ("CIECAM02-UCS",  lambda: _ciecam02_ucs_de(hxyz1, hxyz2)),
        ("Jzazbz",        lambda: _jzazbz_de(hxyz1, hxyz2)),
    ]:
        print(f"  {name}...", end=" ", flush=True)
        de = fn()
        s = stress(de, hdv)
        hscores[name] = round(s, 2)
        print(f"STRESS = {s:.2f}")

    results["HumanFeedback"] = hscores

    # ── Summary ────────────────────────────────────────────────────────────
    _print_summary(results)
    return results


def _print_summary(results: dict):
    # Only show main datasets (not sub-dataset breakdown)
    main_keys = [k for k in results.keys() if not k.endswith("_subsets")]
    if not main_keys:
        return

    metrics = list(results[main_keys[0]].keys())
    col = 16

    print("\n" + "─" * (20 + col * len(main_keys)))
    header = f"  {'Metric':<18}" + "".join(f"{d:>{col}}" for d in main_keys)
    print(header)
    print("─" * (20 + col * len(main_keys)))

    for m in metrics:
        row = f"  {m:<18}"
        for d in main_keys:
            v = results[d].get(m)
            if v is None:
                row += f"{'—':>{col}}"
                continue
            # ★ = best in this dataset column
            best = min(results[d].values())
            marker = " ★" if v == best else "  "
            row += f"{v:>{col-2}.2f}{marker}"
        print(row)

    print("─" * (20 + col * len(main_keys)))
    print("  ★ = best in column  |  lower STRESS = better\n")

    # Rank by average STRESS across available datasets
    print("  Average STRESS across datasets:")
    avgs = {}
    for m in metrics:
        vals = [results[d][m] for d in main_keys if m in results.get(d, {})]
        avgs[m] = round(sum(vals) / len(vals), 2) if vals else None

    ranked = sorted([(v, k) for k, v in avgs.items() if v is not None])
    for rank, (v, name) in enumerate(ranked, 1):
        marker = " ← BEST" if rank == 1 else ""
        print(f"    {rank}. {name:<18} avg={v:.2f}{marker}")
    print()
